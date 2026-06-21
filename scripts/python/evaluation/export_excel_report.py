"""
scripts/python/evaluation/export_excel_report.py
Genera un archivo Excel unificado con pestañas (sheets) que contienen las métricas,
los datos de origen (raw) y las gráficas generadas por Python.

Autor: SISTAC Team
"""

from __future__ import annotations

import csv
from pathlib import Path
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

import sys

# Setup python path to import config.py
_SCRIPTS_DIR = Path(__file__).parent.parent
if str(_SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(_SCRIPTS_DIR))

from config import (
    PROJECT_ROOT,
    TABLES_DIR,
    FIGURES_DIR as CONFIG_FIGURES_DIR,
    GOLD_STANDARD_DIR,
)

FIGURES_DIR = CONFIG_FIGURES_DIR / "cap5"

def generate_excel_report():
    import os
    llm_provider = os.getenv("LLM_PROVIDER", "anthropic").lower()
    provider_tables_dir = TABLES_DIR / llm_provider
    provider_tables_dir.mkdir(parents=True, exist_ok=True)
    excel_path = provider_tables_dir / "reporte_completo_sistac.xlsx"
    
    # Crear un nuevo libro Excel
    wb = openpyxl.Workbook()
    # Eliminar la pestaña por defecto
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # Estilos de celda
    font_title = Font(name="Calibri", size=14, bold=True, color="1F4E79")
    font_desc = Font(name="Calibri", size=11, italic=True, color="595959")
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_body = Font(name="Calibri", size=11)
    fill_header = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    fill_zebra = PatternFill(start_color="F2F5F8", end_color="F2F5F8", fill_type="solid")
    border_thin = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    
    def format_sheet_common(ws, title, desc):
        ws.views.sheetView[0].showGridLines = True
        ws.cell(row=1, column=1, value=title).font = font_title
        ws.cell(row=2, column=1, value=desc).font = font_desc
        ws.row_dimensions[1].height = 25
        ws.row_dimensions[2].height = 20
        ws.row_dimensions[3].height = 10 # Espaciador
        
    # ── PESTAÑA 1: Resumen y Gráficas ──────────────────────────────────────────
    ws_resumen = wb.create_sheet(title="Resumen y Gráficas")
    format_sheet_common(
        ws_resumen, 
        "Reporte Ejecutivo de Resultados - SISTAC", 
        "Visualización de las hipótesis del TFE y diagramas de arquitectura del sistema."
    )
    
    # Agregar breves descripciones del experimento C0-C3
    ws_resumen.cell(row=4, column=1, value="Estructura Factorial del Experimento:").font = Font(name="Calibri", size=11, bold=True)
    exp_details = [
        "• C0: Baseline de screening manual realizado por evaluadores de Recursos Humanos.",
        "• C1: Screening automatizado mediante Large Language Model (LLM) puro (sin RAG).",
        "• C2: Screening automatizado mediante RAG integrado con Azure AI Search (sin anonimizar).",
        "• C3: Screening con anonimización de datos sensibles (PII) + RAG (Azure AI Search + Presidio)."
    ]
    for idx, detail in enumerate(exp_details):
        ws_resumen.cell(row=5 + idx, column=1, value=detail).font = font_body
        
    # Embedder de imágenes de matplotlib (si existen)
    images_to_embed = [
        ("fig5_1_arquitectura_general.png", "A11", "Arquitectura general C0-C3"),
        ("fig5_5_scoring_engine.png", "A32", "Motor de puntuación y dimensiones evaluadas")
    ]
    
    for img_name, cell_loc, label in images_to_embed:
        img_path = FIGURES_DIR / img_name
        if img_path.exists():
            try:
                # Escribir etiqueta del gráfico
                row_label = int(cell_loc[1:]) - 1
                ws_resumen.cell(row=row_label, column=1, value=f"[Gráfico] {label}").font = Font(name="Calibri", size=11, bold=True, color="2E75B6")
                
                # Cargar y redimensionar/insertar imagen
                img = openpyxl.drawing.image.Image(str(img_path))
                # Reducir un poco el tamaño para que se vea bien en Excel
                img.width = img.width * 0.55
                img.height = img.height * 0.55
                ws_resumen.add_image(img, cell_loc)
            except Exception as e:
                print(f"[WARN] No se pudo incrustar imagen {img_name}: {e}")
                
    # ── PESTAÑAS DE HIPÓTESIS (H1, H2, H3) ──────────────────────────────────────
    hypotheses = [
        (
            "H1 - Eficiencia", 
            "Métricas de Eficiencia (H1)", 
            "Objetivo: Comparar el tiempo de procesamiento por candidato en segundos (Mann-Whitney U y Speedup).", 
            provider_tables_dir / "tab_resultados_h1.csv"
        ),
        (
            "H2 - Eficacia", 
            "Métricas de Eficacia (H2)", 
            "Objetivo: Evaluar la precisión global F1-Score y AUC-ROC frente al Gold Standard (umbral F1 >= 0.85).", 
            provider_tables_dir / "tab_resultados_h2.csv"
        ),
        (
            "H3 - Equidad", 
            "Métricas de Equidad (H3)", 
            "Objetivo: Verificar el impacto dispar (DIR) y diferencia de paridad estadística (SPD) por género (umbral DIR >= 0.80).", 
            provider_tables_dir / "tab_resultados_h3.csv"
        ),
        (
            "RAGAS - Calidad", 
            "Métricas RAGAS de Calidad RAG", 
            "Objetivo: Analizar la fidelidad y precisión del contexto recuperado de Azure Search.", 
            provider_tables_dir / "tab_ragas_c2.csv"
        )
    ]
    
    for title_sheet, title_text, desc, csv_path in hypotheses:
        if csv_path.exists():
            ws = wb.create_sheet(title=title_sheet)
            format_sheet_common(ws, title_text, desc)
            
            # Cargar CSV usando pandas
            df = pd.read_csv(csv_path)
            
            # Escribir cabeceras en fila 4
            for col_idx, col_name in enumerate(df.columns, 1):
                cell = ws.cell(row=4, column=col_idx, value=col_name)
                cell.font = font_header
                cell.fill = fill_header
                cell.alignment = align_center
                cell.border = border_thin
                
            # Escribir datos
            for row_idx, row_data in enumerate(df.values, 5):
                ws.row_dimensions[row_idx].height = 20
                for col_idx, val in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    cell.font = font_body
                    cell.border = border_thin
                    cell.alignment = align_center if col_idx > 1 else align_left
                    # Zebra color
                    if row_idx % 2 == 0:
                        cell.fill = fill_zebra
                        
            # Ajustar anchos de columna
            for col in ws.columns:
                max_len = 0
                for cell in col:
                    if cell.row >= 4 and cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                col_letter = get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max(max_len + 4, 15)
                
    # ── PESTAÑA FINAL: Datos de Origen ─────────────────────────────────────────
    gt_path = GOLD_STANDARD_DIR / "ground_truth.csv"
    c0_path = GOLD_STANDARD_DIR / "c0_times.csv"
    
    if gt_path.exists():
        ws_source = wb.create_sheet(title="Datos de Origen (Raw)")
        format_sheet_common(
            ws_source, 
            "Gold Standard y Tiempos de Referencia", 
            "Datos originales de ground truth (etiquetas deseadas) y tiempos registrados manualmente por evaluadores (C0)."
        )
        
        # Cargar datos
        df_gt = pd.read_csv(gt_path)
        
        # Escribir cabeceras
        for col_idx, col_name in enumerate(df_gt.columns, 1):
            cell = ws_source.cell(row=4, column=col_idx, value=col_name)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
            cell.border = border_thin
            
        # Escribir datos (primeros 100 por simplicidad y no hacer el archivo enorme)
        limit_rows = min(len(df_gt), 100)
        for row_idx, row_data in enumerate(df_gt.values[:limit_rows], 5):
            ws_source.row_dimensions[row_idx].height = 18
            for col_idx, val in enumerate(row_data, 1):
                cell = ws_source.cell(row=row_idx, column=col_idx, value=val)
                cell.font = font_body
                cell.border = border_thin
                cell.alignment = align_center
                if row_idx % 2 == 0:
                    cell.fill = fill_zebra
                    
        # Nota final de truncado si aplica
        if len(df_gt) > limit_rows:
            ws_source.cell(row=limit_rows + 5, column=1, value=f"... y {len(df_gt) - limit_rows} filas adicionales en ground_truth.csv original.").font = font_desc
            
        # Ajustar anchos
        for col in ws_source.columns:
            max_len = 0
            for cell in col:
                if cell.row >= 4 and cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            col_letter = get_column_letter(col[0].column)
            ws_source.column_dimensions[col_letter].width = max(max_len + 4, 15)
            
    # Guardar archivo Excel
    wb.save(str(excel_path))
    print(f"  [OK] Reporte Excel unificado creado en: {excel_path.name}")

if __name__ == "__main__":
    generate_excel_report()
